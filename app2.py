# =====================================
# Streamlit Web App: 模拟Project：人事用合同记录表自动审核（四输出表版 + 漏填检查 + 驻店客户版）
# =====================================

import streamlit as st
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows # <--- 添加或确保这一行存在
from io import BytesIO

def normalize_contract_key(series: pd.Series) -> pd.Series:
    """
    对合同号 Series 进行标准化处理，用于安全的 pd.merge 操作。
    """
    # 1. 确保是字符串类型，同时处理缺失值（如果存在）
    s = series.astype(str)
    
    # 2. 移除常见的浮点数残留（以防原始数据错误输入）
    s = s.str.replace(r"\.0$", "", regex=True) 
    
    # 3. 核心：移除首尾空格（处理最常见的导入错误）
    s = s.str.strip()
    
    # 4. 统一转换为大写（处理大小写不一致问题，如 'pazl' vs 'PAZL'）
    s = s.str.upper() 
    
    # 5. 处理全角/半角差异（将常见的全角连接符转为半角）
    s = s.str.replace('－', '-', regex=False) # 全角连接符转半角
    
    # 6. 处理其他可能的空白字符（例如 tabs, 换行符等）
    s = s.str.replace(r'\s+', '', regex=True)
    
    return s

# =====================================
# 🏁 应用标题与说明
# =====================================
st.title("📊 模拟人事用薪资计算表自动审核系统-1重卡")

st.image("image/app1(1).png")

# =====================================
# 📂 上传文件区：要求上传 4 个 xlsx 文件
# =====================================
uploaded_files = st.file_uploader(
    "请上传文件名中包含以下字段的文件：月重卡、放款明细、字段、二次明细。最后誊写，需检的表为文件名包含‘月重卡’字段的表。",
    type="xlsx",
    accept_multiple_files=True
)

if not uploaded_files or len(uploaded_files) < 4:
    st.warning("⚠️ 请上传所有 4 个文件后继续")
    st.stop()
else:
    st.success("✅ 文件上传完成")

# =====================================
# 🧰 工具函数区（文件定位、列名模糊匹配、日期/数值处理）
# =====================================

# 按关键字查找文件（文件名包含关键字）
def find_file(files_list, keyword):
    for f in files_list:
        if keyword in f.name:
            return f
    raise FileNotFoundError(f"❌ 未找到包含关键词「{keyword}」的文件")

# 统一列名格式（去空格、转小写）
def normalize_colname(c): return str(c).strip().lower()

# 按关键字匹配列名（支持 exact 精确匹配与模糊匹配）
def find_col(df, keyword, exact=False):
    key = keyword.strip().lower()
    for col in df.columns:
        cname = normalize_colname(col)
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None

# 查找sheet（sheet名包含关键字即可）
def find_sheet(xls, keyword):
    for s in xls.sheet_names:
        if keyword in s:
            return s
    raise ValueError(f"❌ 未找到包含关键词「{keyword}」的sheet")

# 统一数值解析（去逗号、转float、处理百分号）
def normalize_num(val):
    if pd.isna(val): return None
    s = str(val).replace(",", "").strip()
    if s in ["", "-", "nan"]: return None
    try:
        if "%" in s: return float(s.replace("%", "")) / 100
        return float(s)
    except ValueError:
        return s

# 日期匹配（年/月/日完全一致）
def same_date_ymd(a, b):
    try:
        da = pd.to_datetime(a, errors='coerce')
        db = pd.to_datetime(b, errors='coerce')
        if pd.isna(da) or pd.isna(db): return False
        return (da.year, da.month, da.day) == (db.year, db.month, db.day)
    except Exception:
        return False
def prepare_ref_df(ref_df, mapping, prefix):
    # --- 修正开始 ---
    
    # 1. 找到参考表(ref_df)中的“合同”列
    # 我们使用 find_col，这才是正确的做法
    contract_col = find_col(ref_df, "合同") 
    
    # 2. 如果在 ref_df 中找不到合同列，则无法继续
    if not contract_col:
        st.warning(f"⚠️ 在 {prefix} 参考表中未找到'合同'列，跳过此数据源。")
        return pd.DataFrame(columns=['__KEY__']) # 返回一个空的带key的df
        
    std_df = pd.DataFrame()
    
    # 3. VVVV 插入归一化函数 VVVV
    # 使用找到的 contract_col 来应用归一化
    std_df['__KEY__'] = normalize_contract_key(ref_df[contract_col])
    # ^^^^ 插入归一化函数 ^^^^
    
    # --- 修正结束 ---
    # 4. 提取并重命名所有需要的字段 (【已修改】增加租赁期限*12的逻辑)
    for main_kw, ref_kw in mapping.items():
        # 城市经理需要精确匹配
        exact = (main_kw == "城市经理") 
        ref_col_name = find_col(ref_df, ref_kw, exact=exact)
        
        if ref_col_name:
            # --- VVVV (这是新添加的逻辑) VVVV ---
            # 获取原始数据 Series
            s_ref_raw = ref_df[ref_col_name]
            
            # 检查是否是 'fk' 表的 '租赁期限'
            if prefix == 'fk' and main_kw == '租赁期限':
                # 应用转换：年 -> 月
                s_ref_transformed = pd.to_numeric(s_ref_raw, errors='coerce') * 12
                std_df[f'ref_{prefix}_{main_kw}'] = s_ref_transformed
            else:
                # 无转换，直接赋值
                std_df[f'ref_{prefix}_{main_kw}'] = s_ref_raw
            # --- ^^^^ (新逻辑结束) ^^^^ ---
        else:
            st.warning(f"⚠️ 在 {prefix} 参考表中未找到列 (main: '{main_kw}', ref: '{ref_kw}')")

    # 5. 效仿原始逻辑：只取第一个匹配项 (这部分逻辑保持不变)
    std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
    return std_df

def compare_series_vec(s_main, s_ref, main_kw):
    """
    向量化比较两个Series，复刻原始的 compare_fields_and_mark 逻辑。
    返回一个布尔Series，True表示存在差异。
    (V2：增加对 merge 失败 (NaN) 的静默跳过)
    """
    
    # 0. 识别真正的 "Merge 失败" (s_ref 是物理 NaN)
    #    我们必须在 s_ref 被 astype(str) 污染前执行此操作
    merge_failed_mask = s_ref.isna() 

    # 1. 预处理：处理空值。原始逻辑：同为NaN/空则认为一致。
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    
    # 两者都为空（NaN, "", "None"等），不算错误
    both_are_na = main_is_na & ref_is_na
    
    # 2. 日期字段比较
    if any(k in main_kw for k in ["日期", "时间"]):
        d_main = pd.to_datetime(s_main, errors='coerce')
        d_ref = pd.to_datetime(s_ref, errors='coerce')
        
        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main.dt.date != d_ref.dt.date)
        
        errors = valid_dates_mask & date_diff_mask
    
    # 3. 数值/文本比较
    else:
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)
        
        main_is_na_norm = pd.isna(s_main_norm) | (s_main_norm.astype(str).str.strip().isin(["", "nan", "None"]))
        ref_is_na_norm = pd.isna(s_ref_norm) | (s_ref_norm.astype(str).str.strip().isin(["", "nan", "None"]))
        both_are_na_norm = main_is_na_norm & ref_is_na_norm

        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref
        
        errors = pd.Series(False, index=s_main.index)

        # 3a. 数值比较
        if both_are_num.any():
            num_main = s_main_norm[both_are_num].fillna(0) # fillna(0) for safety
            num_ref = s_ref_norm[both_are_num].fillna(0)
            diff = (num_main - num_ref).abs()
            
            # --- VVVV (这是修改后的逻辑) VVVV ---
            if main_kw == "保证金比例":
                num_errors = (diff > 0.00500001) # 保证金比例容错
            
            elif "租赁期限" in main_kw: # 匹配 "租赁期限" 和 "租赁期限月"
                # 忽略小于 1.0 个月的差距 (即：差异 >= 1.0 才算错误)
                num_errors = (diff >= 1.0) 
            
            else:
                # 其他数值字段，使用标准微小容错
                num_errors = (diff > 1e-6)
            # --- ^^^^ (修改结束) ^^^^ ---
            
            errors.loc[both_are_num] = num_errors

        # 3b. 文本比较
        not_num_mask = ~both_are_num
        if not_num_mask.any():
            str_main = s_main_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            str_ref = s_ref_norm[not_num_mask].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
            
            str_errors = (str_main != str_ref)
            errors.loc[not_num_mask] = str_errors
            
        # 排除掉那些两者皆为空的情况
        errors = errors & ~both_are_na_norm
        # return errors # (这是旧的返回)

    # 4. === 最终错误逻辑 ===
    
    # a. 排除 "两者皆为空" 的情况 (原始逻辑)
    final_errors = errors & ~both_are_na
    
    # b. 排除 "Merge 失败" 的情况 (复刻 iterrows 的 'if ref_rows.empty: return 0')
    #    条件:
    #    1. merge_failed_mask 为 True (s_ref 是物理 NaN)
    #    2. main_is_na 为 False (s_main 不是空的)
    #    如果 (1) 和 (2) 都成立，说明这是一个 "lookup failure"，我们必须忽略它
    
    lookup_failure_mask = merge_failed_mask & ~main_is_na
    
    final_errors = final_errors & ~lookup_failure_mask
    
    return final_errors

# =====================================
# 🧮 单sheet检查函数 (向量化版)
# =====================================
def check_one_sheet(sheet_keyword, main_file, ref_dfs_std_dict):
    start_time = time.time()
    xls_main = pd.ExcelFile(main_file)

    # 查找目标sheet
    try:
        target_sheet = find_sheet(xls_main, sheet_keyword)
    except ValueError:
        st.warning(f"⚠️ 未找到包含「{sheet_keyword}」的sheet，跳过。")
        return 0, None, 0, set()

    # 1. 读取目标sheet（第二行为表头）
    try:
        main_df = pd.read_excel(xls_main, sheet_name=target_sheet, header=1)
    except Exception as e:
        st.error(f"❌ 读取「{sheet_keyword}」时出错: {e}")
        return 0, None, 0, set()
        
    if main_df.empty:
        st.warning(f"⚠️ 「{sheet_keyword}」为空，跳过。")
        return 0, None, 0, set()

    # 2. 查找合同号列
    global contract_col_main
    contract_col_main = find_col(main_df, "合同")
    if not contract_col_main:
        st.error(f"❌ 在「{sheet_keyword}」中未找到合同列。")
        return 0, None, 0, set()

    # 3. 创建临时输出文件 (保留原始表头空行)
    output_path = f"月重卡_{sheet_keyword}_审核标注版.xlsx"
    empty_row = pd.DataFrame([[""] * len(main_df.columns)], columns=main_df.columns)
    # 注意：这里我们保存的是原始main_df
    pd.concat([empty_row, main_df], ignore_index=True).to_excel(output_path, index=False)

    # 打开Excel用于写入标注
    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

   # 4. 准备主表用于合并
    # (注意：contract_col_main 已经在第 236 行被正确找到了，我们不需要再找了)
    
    # 存储原始索引，用于 openpyxl 定位
    main_df['__ROW_IDX__'] = main_df.index
    
    # VVVV 插入归一化函数 VVVV
    # 创建标准合并Key
    main_df['__KEY__'] = normalize_contract_key(main_df[contract_col_main])
    # ^^^^ 插入归一化函数 ^^^^
    
    # 获取本表所有合同号（用于统计等）
    contracts_seen = set(main_df['__KEY__'].dropna())

    # 5. 一次性合并所有参考数据
    merged_df = main_df.copy()
    for prefix, std_df in ref_dfs_std_dict.items():
        if not std_df.empty:
            merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')
    
    total_errors = 0
    skip_city_manager = [0]
    errors_locations = set() # 存储 (row_idx, col_name)
    row_has_error = pd.Series(False, index=merged_df.index) # 标记哪一行有错误

    # 添加Streamlit进度条
    progress = st.progress(0)
    status = st.empty()

    # 6. === 遍历字段进行向量化比对 ===
    mappings_all = {
        'fk': (mapping_fk, ref_dfs_std_dict['fk']),
        'zd': (mapping_zd, ref_dfs_std_dict['zd']),
        'ec': (mapping_ec, ref_dfs_std_dict['ec'])
    }
    
    total_comparisons = sum(len(m[0]) for m in mappings_all.values())
    current_comparison = 0

    for prefix, (mapping, std_df) in mappings_all.items():
        if std_df.empty:
            current_comparison += len(mapping) # 跳过空表
            continue
            
        for main_kw, ref_kw in mapping.items():
            current_comparison += 1
            status.text(f"检查「{sheet_keyword}」: {prefix} - {main_kw}...")
            
            # 关键：在原始 main_df 中找到列名
            exact = (main_kw == "城市经理")
            main_col = find_col(main_df, main_kw, exact=exact)
            
            # 参考列的列名是我们在 prepare_ref_df 中标准化的
            ref_col = f'ref_{prefix}_{main_kw}'

            if not main_col or ref_col not in merged_df.columns:
                continue # 跳过不存在的列

            s_main = merged_df[main_col]
            s_ref = merged_df[ref_col]

            # 处理 "城市经理" 跳过逻辑
            skip_mask = pd.Series(False, index=merged_df.index)
            if main_kw == "城市经理":
                na_strings = ["", "-", "nan", "none", "null"]
                # 检查参考列是否为空
                skip_mask = pd.isna(s_ref) | s_ref.astype(str).str.strip().isin(na_strings)
                skip_city_manager[0] += skip_mask.sum()
            
            # 7. 获取向量化比较结果
            errors_mask = compare_series_vec(s_main, s_ref, main_kw)
            
            # 应用跳过逻辑：如果 skip_mask 为 True，则不算错误
            final_errors_mask = errors_mask & ~skip_mask
            
            if final_errors_mask.any():
                total_errors += final_errors_mask.sum()
                row_has_error |= final_errors_mask
                
                # 8. 存储错误位置 (使用 __ROW_IDX__ 和 原始 main_col 名称)
                bad_indices = merged_df[final_errors_mask]['__ROW_IDX__']
                for idx in bad_indices:
                    errors_locations.add((idx, main_col))
                    
            progress.progress(current_comparison / total_comparisons)

    status.text(f"「{sheet_keyword}」比对完成，正在生成标注文件...")

    # 9. === 遍历错误进行Excel标注 ===
    # (这比遍历所有单元格快得多)
    
    # 获取原始列名 (去掉我们添加的辅助列)
    original_cols_list = list(main_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
    # 创建列名到Excel列索引(1-based)的映射
    col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}

    # 标红错误单元格
    for (row_idx, col_name) in errors_locations:
        if col_name in col_name_to_idx:
            # +3: (1-based index) + (1 for header) + (1 for empty row)
            ws.cell(row_idx + 3, col_name_to_idx[col_name]).fill = red_fill

    # 标黄有错误的合同号
    if contract_col_main in col_name_to_idx:
        contract_col_excel_idx = col_name_to_idx[contract_col_main]
        # 找到所有出错的原始行号
        error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
        for row_idx in error_row_indices:
            ws.cell(row_idx + 3, contract_col_excel_idx).fill = yellow_fill

    # 10. 导出检查结果
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.download_button(
        label=f"📥 下载 {sheet_keyword}审核标注版",
        data=output,
        file_name=f"记录表_{sheet_keyword}_审核标注版.xlsx",
        key=f"download_{sheet_keyword}" # 增加key避免streamlit重跑问题
    )

    # 11. (新) 导出仅含错误行的文件 (带标红)
    if row_has_error.any():
        try:
            # 1. 获取仅含错误行的 DataFrame (只保留原始列)
            #    (original_cols_list 已在第 365 行定义)
            df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()
            
            # 2. 关键：创建 "原始行索引" 到 "新Excel行号" 的映射
            #    我们获取所有出错行的 __ROW_IDX__
            original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']
            
            #    创建映射: { 原始索引 : 新的Excel行号 }
            #    (enumerate start=2, 因为 Excel 行 1 是表头, 数据从行 2 开始)
            original_idx_to_new_excel_row = {
                original_idx: new_row_num 
                for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
            }

            # 3. 创建一个新的工作簿(Workbook)
            wb_errors = Workbook()
            ws_errors = wb_errors.active
            
            # 4. 使用 dataframe_to_rows 快速写入数据
            for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                ws_errors.append(r)
                
            # 5. 遍历主错误列表(errors_locations)，进行标红
            #    (col_name_to_idx 和 red_fill 已在前面定义)
            for (original_row_idx, col_name) in errors_locations:
                
                # 检查这个错误是否在我们 "仅错误行" 的映射中
                if original_row_idx in original_idx_to_new_excel_row:
                    
                    # 获取它在新Excel文件中的行号
                    new_row = original_idx_to_new_excel_row[original_row_idx]
                    
                    # 获取列号
                    if col_name in col_name_to_idx:
                        new_col = col_name_to_idx[col_name]
                        
                        # 应用标红
                        ws_errors.cell(row=new_row, column=new_col).fill = red_fill
            
            # 6. 保存到 BytesIO
            output_errors_only = BytesIO()
            wb_errors.save(output_errors_only)
            output_errors_only.seek(0)
            
            # 7. 创建下载按钮
            st.download_button(
                label=f"📥 下载 {sheet_keyword} (仅含错误行, 带标红)", # 更新了标签
                data=output_errors_only,
                file_name=f"记录表_{sheet_keyword}_仅错误行_标红.xlsx", # 更新了文件名
                key=f"download_{sheet_keyword}_errors_only" # Key 保持不变
            )
        except Exception as e:
            st.error(f"❌ 生成“仅错误行”文件时出错: {e}")
    
    elapsed = time.time() - start_time
    st.success(f"✅ {sheet_keyword} 检查完成，共 {total_errors} 处错误，用时 {elapsed:.2f} 秒。")
    return total_errors, elapsed, skip_city_manager[0], contracts_seen

# =====================================
# 📖 文件读取：按关键字识别五份文件
# =====================================
main_file = find_file(uploaded_files, "月重卡")
fk_file = find_file(uploaded_files, "放款明细")
zd_file = find_file(uploaded_files, "字段")
ec_file = find_file(uploaded_files, "二次明细")

# 各文件sheet读取（模糊匹配sheet名）
fk_df = pd.read_excel(pd.ExcelFile(fk_file), sheet_name=find_sheet(pd.ExcelFile(fk_file), "威田"))
zd_df = pd.read_excel(pd.ExcelFile(zd_file), sheet_name=find_sheet(pd.ExcelFile(zd_file), "重卡"))
ec_df = pd.read_excel(ec_file)

# 合同列定位
contract_col_fk = find_col(fk_df, "合同")
contract_col_zd = find_col(zd_df, "合同")
contract_col_ec = find_col(ec_df, "合同")

# 对照字段映射表
# --- VVVV (这是新的，修正的) VVVV ---
# 格式: {"记录表(主表)的列名": "放款明细(参考表)的列名关键字"}
mapping_fk = {
    # 这3个是正确的
    "授信方": "授信方",     # "授信方" in "授信方"
    "租赁本金": "租赁本金",  
    "租赁期限": "租赁期限",
    
    # 这4个是修正的
    "挂车台数": "挂车数量",     # "挂车数量" in "挂车数量"
    "起租收益率": "XIRR"      # 假设 "费率" 是您想要的 "收益率"。如果不是，请修改为 "XIRR" 或其他
}
# --- ^^^^ (修正结束) ^^^^ ---


mapping_zd = {"保证金比例": "保证金比例_2", "项目提报人": "提报", "起租时间": "起租日_商", "客户经理": "客户经理_资产", "所属省区": "区域", "主车台数": "主车台数", "城市经理": "城市经理"}
mapping_ec = {"二次时间": "出本流程时间"}

# =====================================
# 🚀 (新) 预处理所有参考表
# =====================================
st.info("ℹ️ 正在预处理参考数据...")

# (fk_df, zd_df, ec_df, zk_df 必须已经加载)
fk_std = prepare_ref_df(fk_df, mapping_fk, 'fk')
zd_std = prepare_ref_df(zd_df, mapping_zd, 'zd')
ec_std = prepare_ref_df(ec_df, mapping_ec, 'ec')

# 将所有预处理过的DF存入字典，传递给检查函数
ref_dfs_std_dict = {
    'fk': fk_std,
    'zd': zd_std,
    'ec': ec_std
}
st.success("✅ 参考数据预处理完成。")

# =====================================
# 🧾 多sheet循环 + 驻店客户表
# =====================================
sheet_keywords = ["二次", "部分担保", "随州", "驻店客户"]
total_all = elapsed_all = skip_total = 0
contracts_seen_all_sheets = set()

# 循环处理四张sheet (调用新函数)
for kw in sheet_keywords:
    # 将 main_file 和 ref_dfs_std_dict 传递进去
    count, used, skipped, seen = check_one_sheet(kw, main_file, ref_dfs_std_dict)
    
    total_all += count
    elapsed_all += used or 0
    skip_total += skipped
    contracts_seen_all_sheets.update(seen)

st.success(f"🎯 全部审核完成，共 {total_all} 处错误，总耗时 {elapsed_all:.2f} 秒。")

# =====================================
# 🕵️ 漏填检查：跳过“是否车管家=是”与“提成类型=联合租赁/驻店”
# =====================================
field_contracts = zd_df[contract_col_zd].dropna().astype(str).str.strip()
col_car_manager = find_col(zd_df, "是否车管家", exact=True)
col_bonus_type = find_col(zd_df, "提成类型", exact=True)

missing_contracts_mask = (~field_contracts.isin(contracts_seen_all_sheets))

# 跳过“车管家=是”
if col_car_manager:
    missing_contracts_mask &= ~(zd_df[col_car_manager].astype(str).str.strip().str.lower() == "是")
# 跳过“联合租赁/驻店”
if col_bonus_type:
    missing_contracts_mask &= ~(
        zd_df[col_bonus_type].astype(str).str.strip().isin(["联合租赁", "驻店"])
    )

# 标记漏填
zd_df_missing = zd_df.copy()
zd_df_missing["漏填检查"] = ""
zd_df_missing.loc[missing_contracts_mask, "漏填检查"] = "❗ 漏填"
漏填合同数 = zd_df_missing["漏填检查"].eq("❗ 漏填").sum()
st.warning(f"⚠️ 共发现 {漏填合同数} 个合同在记录表中未出现（已排除车管家、联合租赁、驻店）")

# =====================================
# 📤 导出字段表（含漏填标注 + 仅漏填版）
# =====================================
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 全字段表（含漏填标注）
wb = Workbook()
ws = wb.active
for c_idx, c in enumerate(zd_df_missing.columns, 1): ws.cell(1, c_idx, c)
for r_idx, row in enumerate(zd_df_missing.itertuples(index=False), 2):
    for c_idx, v in enumerate(row, 1):
        ws.cell(r_idx, c_idx, v)
        if zd_df_missing.columns[c_idx-1] == "漏填检查" and v == "❗ 漏填":
            ws.cell(r_idx, c_idx).fill = yellow_fill

output_all = BytesIO()
wb.save(output_all)
output_all.seek(0)
st.download_button("📥 下载字段表漏填标注版", output_all, "字段表_漏填标注版.xlsx")

# 仅漏填合同
zd_df_only_missing = zd_df_missing[zd_df_missing["漏填检查"] == "❗ 漏填"].copy()
if not zd_df_only_missing.empty:
    wb2 = Workbook()
    ws2 = wb2.active
    for c_idx, c in enumerate(zd_df_only_missing.columns, 1): ws2.cell(1, c_idx, c)
    for r_idx, row in enumerate(zd_df_only_missing.itertuples(index=False), 2):
        for c_idx, v in enumerate(row, 1):
            ws2.cell(r_idx, c_idx, v)
            if zd_df_only_missing.columns[c_idx-1] == "漏填检查" and v == "❗ 漏填":
                ws2.cell(r_idx, c_idx).fill = yellow_fill
    out2 = BytesIO()
    wb2.save(out2)
    out2.seek(0)
    st.download_button("📥 下载仅漏填字段表", out2, "字段表_仅漏填.xlsx")

# =====================================
# ✅ 结束提示
# =====================================
st.success("✅ 所有检查、标注与导出完成！")
